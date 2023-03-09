# 个人开发&学习
# 开发人员：lucky
# 开发时间：2022/12/8 21:38
# 名言名句：万般皆下品惟有读书高

# 爬虫案列
'''
import requests
import json

url = 'https://club.jd.com/comment/productPageComments.action?callback=fetchJSON_comment98&productId=100006775445&score=0&sortType=5&page=0&pageSize=10&isShadowSku=0&fold=1'
headers = {
    'accept': ' */*',
    'cache-control': 'max-age=0',
    'content-type': 'text/html;charset=GBK',
    'referer': 'https://item.jd.com/',
    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/108.0.0.0 Safari/537.36'
}
frp = requests.get(url, headers=headers)
conent = frp.text
ret = conent.replace('fetchJSON_comment98(', '').replace(');', '')
json_data = json.loads(ret)
comments = json_data['comments']
for item in comments:
    color = item['productColor']
    idr = item['id']
    print(color)
    print(idr)
'''

import requests
import json
import openpyxl

wk = openpyxl.Workbook()
sheet = wk.create_sheet()
for item in range(0, 2):
    url = 'https://club.jd.com/comment/productPageComments.action?callback=fetchJSON_comment98&productId=100006775445&score=0&sortType=5&page=0&pageSize=10&isShadowSku=0&fold=1'
    frp = requests.get(url)
    conent = frp.text
    ret = conent.replace('fetchJSON_comment98(', '').replace(');', '')
    json_data = json.loads(ret)
    comments = json_data['comments']
    for i in comments:
        color = i['productColor']
        idr = i['id']
        sheet.append([color, idr])
        wk.save('data/lucky-18374781741.xlsx')

# Excel数据可视化分析
import openpyxl
import matplotlib.pyplot as plt

wk = openpyxl.load_workbook('data/lucky-18374781741.xlsx')
sheet = wk['Sheet1']
colors = []
idrs = []
for i in range(1, 21):
    colors.append(sheet['A' + str(i).value])
    idrs.append(sheet['B' + str(i).value])
color_class = set(colors)
count = len(colors)
color_percent = []
for clr in color_class:
    color_percent.append(colors.count(clr) / count)

plt.pie(x=color_percent, labels=color_class, autopct='')
plt.rcParams['font.sans-serif'] = ['SimHei']
plt.legend()
plt.savefig('data/lucky-18374781741.png')
